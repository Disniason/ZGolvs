import os
import json
import openpyxl
from io import BytesIO

from database.sql import sql
from database.db import async_database as db



class vocabulary:

    @classmethod
    async def add_vocabulary_table(cls, table_name: str):
        if not table_name or len(table_name.strip()) < 1:
            return {"code": 0, "msg": "单词表名称不能为空"}
        clean_name = table_name.strip()

        # 第一步：先查重
        check_sql = sql.check_name_vocabulary_exists_sql()
        rows = await db.execute_sql(check_sql, (clean_name,))
        count = rows[0][0]
        if count > 0:
            return {"code": 0, "msg": "该单词表名称已存在，请更换名称"}

        # 名称不存在再执行建表+插入
        sql_str = sql.create_word_table_insert_vocabulary_sql(clean_name)
        try:
            affect_rows = await db.execute_sql(sql_str, (clean_name, clean_name, clean_name))
        except Exception as e:
            print("数据库异常", repr(e))
            return {"code": 0, "msg": f"创建失败：{str(e)}"}

        if affect_rows >= 0:
            return {"code": 1, "msg": "单词表创建成功"}
        else:
            return {"code": 0, "msg": "创建执行无生效行数"}


    @classmethod
    async def get_all_vocabulary_tables(cls):
        sql_str = sql.get_all_vocabulary_list_sql()
        rows = await db.execute_sql(sql_str)
        table_list = []
        for row in rows:
            table_list.append({
                "vocabularyId": row[0],
                "vocabularyName": row[1],
                "createTime": row[2]
            })
        return {"code": 1, "data": table_list}
    

    @classmethod
    async def search_vocabulary_tables(cls, keyword: str):
        all_tables_res = await cls.get_all_vocabulary_tables()
        all_tables = all_tables_res.get("data", [])

        # 内存中过滤（不区分大小写，不拼接SQL）
        if not keyword or keyword.strip() == "":
            return all_tables_res

        filter_list = [
            table for table in all_tables
            if keyword.lower() in table["vocabularyName"].lower()
        ]
        # 保持和原接口一致的返回结构
        return {"code": 1, "data": filter_list}


    @classmethod
    async def delete_vocabulary_table(cls, table_name: str):
        if not table_name or len(table_name.strip()) < 1:
            return {"code": 0, "msg": "单词表名称不能为空"}
        clean_name = table_name.strip()
        sql_str = sql.delete_word_table_sql(clean_name)
        affect_rows = await db.execute_sql(sql_str, (clean_name,))

        if affect_rows >= 0:
            return {"code": 1, "msg": "单词表及内部单词数据已全部删除"}
        else:
            return {"code": 0, "msg": "删除单词表操作失败"}


    @classmethod
    async def import_word_from_json(cls, table_name: str, json_file_bytes):
        """
        # 每批插入条数，可微调300~800，数值越小越稳定
        BATCH_SIZE = 500

        # 1.校验单词表是否存在
        check_table_sql = sql.check_name_vocabulary_exists_sql()
        table_exist_rows = await db.execute_sql(check_table_sql, (table_name,))
        if table_exist_rows[0][0] == 0:
            return {"code": 0, "msg": f"单词表【{table_name}】不存在"}

        # 2.解析JSON
        try:
            json_text = json_file_bytes.decode("utf-8")
            json_data = json.loads(json_text)
        except Exception as e:
            return {"code": 0, "msg": f"JSON文件解析失败：{str(e)}"}

        valid_word_list = []
        skip_count = 0

        # 3.遍历清洗数据
        for key, word_item in json_data.items():
            if not isinstance(word_item, dict) or len(word_item) == 0:
                skip_count += 1
                continue
            word = word_item.get("word")
            phonetic = word_item.get("phonetic")
            translations_arr = word_item.get("translations")
            # 必填校验
            if not word or not phonetic or not translations_arr or not isinstance(translations_arr, list):
                skip_count += 1
                continue

            # 格式化字段
            translations_str = "；".join(translations_arr)
            exchanges_str = json.dumps(word_item.get("exchanges", {}), ensure_ascii=False)
            examples_arr = word_item.get("examples", [])
            examples_str = "|||".join(examples_arr) if isinstance(examples_arr, list) else ""
            uk_path = word_item.get("uk_audio_path", "")
            us_path = word_item.get("us_audio_path", "")

            valid_word_list.append((
                word, translations_str, exchanges_str, examples_str, phonetic, uk_path, us_path
            ))

        total_valid = len(valid_word_list)
        if total_valid == 0:
            return {"code": 0, "msg": f"无合法可导入单词，跳过数量：{skip_count}"}

        total_insert = 0
        # 4. 拆分小批次循环插入
        try:
            for start in range(0, total_valid, BATCH_SIZE):
                batch_items = valid_word_list[start:start + BATCH_SIZE]
                batch_len = len(batch_items)
                # 生成当前批次SQL
                batch_sql = sql.batch_insert_word_sql(table_name, batch_len)
                # 平铺参数
                flat_params = []
                for item in batch_items:
                    flat_params.extend(item)
                # 执行单批插入
                affect = await db.execute_sql(batch_sql, tuple(flat_params))
                total_insert += affect
        except Exception as e:
            print("批量插入异常详情", repr(e))
            return {
                "code": 0,
                "msg": f"数据库插入中断，已成功写入{total_insert}条，失败原因：{str(e)}"
            }

        return {
            "code": 1,
            "msg": f"导入完成！成功{total_insert}条，跳过非法数据{skip_count}条"
        }
        """
        # 固定长度限制，word主键、音标做长度约束
        MAX_WORD = 500
        MAX_PHONETIC = 500

        # 第一步：校验目标单词表是否真实存在于vocabulary_table库
        check_table_exist_sql = sql.check_name_vocabulary_exists_sql()
        table_check_rows = await db.execute_sql(check_table_exist_sql, (table_name,))
        if table_check_rows[0][0] == 0:
            return {"code": 0, "msg": f"导入失败：单词表【{table_name}】不存在"}

        # 第二步：解析JSON文件
        try:
            json_content = json_file_bytes.decode("utf-8")
            json_dict = json.loads(json_content)
        except UnicodeDecodeError:
            return {"code": 0, "msg": "文件编码错误，请使用UTF-8格式JSON文件"}
        except json.JSONDecodeError as e:
            return {"code": 0, "msg": f"JSON格式解析失败：{str(e)}"}

        valid_insert_data = []
        total_skip = 0

        # 第三步：遍历JSON，过滤清洗合法单词数据
        for key, word_raw_data in json_dict.items():
            # 过滤空字典、非字典无效节点（如示例里 "a": {}）
            if not isinstance(word_raw_data, dict) or len(word_raw_data) == 0:
                total_skip += 1
                continue

            # 提取必填字段
            word = word_raw_data.get("word")
            phonetic = word_raw_data.get("phonetic")
            trans_list = word_raw_data.get("translations")

            # 必填项校验：单词、音标、释义数组缺一不可
            if not word or not phonetic or not trans_list or not isinstance(trans_list, list):
                total_skip += 1
                continue

            # 格式化拼接多值字段
            trans_str = "；".join(trans_list)
            exchange_json_str = json.dumps(word_raw_data.get("exchanges", {}), ensure_ascii=False)
            example_list = word_raw_data.get("examples", [])
            example_str = "|||".join(example_list) if isinstance(example_list, list) else ""

            # 音频路径兜底空字符串，禁止出现None
            uk_audio = word_raw_data.get("uk_audio_path", "")
            us_audio = word_raw_data.get("us_audio_path", "")

            # 对固定长度字段裁切，防止主键超长
            safe_word = word[:MAX_WORD]
            safe_phonetic = phonetic[:MAX_PHONETIC]

            # 组装单条插入参数元组
            insert_params = (
                safe_word,
                trans_str,
                exchange_json_str,
                example_str,
                safe_phonetic,
                uk_audio,
                us_audio
            )
            valid_insert_data.append(insert_params)

        # 无合法可插入数据直接返回
        if len(valid_insert_data) == 0:
            return {
                "code": 0,
                "msg": f"无有效单词可导入，过滤跳过数据总量：{total_skip}"
            }

        # 第四步：获取单条插入SQL语句
        single_insert_sql = sql.insert_single_word_sql(table_name)
        success_count = 0
        fail_count = 0
        fail_detail = []

        # 循环单条逐行插入
        for param_tuple in valid_insert_data:
            try:
                await db.execute_sql(single_insert_sql, param_tuple)
                success_count += 1
            except Exception as err:
                fail_count += 1
                word_name = param_tuple[0]
                fail_detail.append(f"单词【{word_name}】异常：{str(err)}")
                total_skip += 1

        # 组装最终返回提示信息
        msg_text = f"导入完成！成功写入{success_count}条，失败{fail_count}条，原始过滤跳过{total_skip - fail_count}条"
        # 如果存在失败单词，追加前3条失败详情提示
        if len(fail_detail) > 0:
            show_err = "；".join(fail_detail[:3])
            msg_text += f"，失败示例：{show_err}"

        return {
            "code": 1,
            "msg": msg_text
        }


    @classmethod
    async def get_table_word_spell_list(cls, table_name: str):
        # 获取指定单词表所有单词拼写
        # 先校验单词表存在
        check_sql = sql.check_name_vocabulary_exists_sql()
        row = await db.execute_sql(check_sql, (table_name,))
        if row[0][0] == 0:
            return {"code":0, "msg":f"单词表{table_name}不存在"}
        
        get_all_spell_sql = sql.get_table_all_spell_sql(table_name)
        rows = await db.execute_sql(get_all_spell_sql)
        spell_list = [r[0] for r in rows]
        return {
            "code":1,
            "msg":"查询成功",
            "spell_list": spell_list
        }


    @classmethod
    async def get_single_word_info(cls, table_name: str, word: str):
        # 查询单个单词完整详情
        # 校验表存在
        check_sql = sql.check_name_vocabulary_exists_sql()
        row = await db.execute_sql(check_sql, (table_name,))
        if row[0][0] == 0:
            return {"code":0, "msg":f"单词表{table_name}不存在"}
        
        get_word_detail_sql = sql.get_single_word_detail_sql(table_name)
        rows = await db.execute_sql(get_word_detail_sql, (word,))
        if not rows:
            return {"code":0, "msg":"该单词不存在"}
        r = rows[0]
        data = {
            "word": r[0],
            "translations": r[1],
            "exchanges": r[2],
            "examples": r[3],
            "phonetic": r[4],
            "uk_audio_path": r[5],
            "us_audio_path": r[6],
            "createTime": str(r[7])
        }
        return {
            "code":1,
            "data": data
        }
    

    @classmethod
    async def get_audio_file_path(cls, table_name: str, word: str, audio_type: str):
        """
        从数据库读取音频本地路径，校验文件是否存在
        :param table_name: 目标单词表名
        :param word: 单词拼写
        :param audio_type: uk / us
        :return: 成功{code:1, file_path:"xxx"} 失败{code:0, msg:"xxx"}
        """
        # 1.校验单词表存在
        check_table_sql = sql.check_name_vocabulary_exists_sql()
        table_check = await db.execute_sql(check_table_sql, (table_name,))
        if table_check[0][0] == 0:
            return {"code": 0, "msg": f"单词表 {table_name} 不存在"}

        # 2.查询单词完整信息，取出音频路径
        get_word_detail_sql = sql.get_single_word_detail_sql(table_name)
        rows = await db.execute_sql(get_word_detail_sql, (word,))
        if not rows:
            return {"code": 0, "msg": f"单词 {word} 不存在"}
        
        row = rows[0]
        uk_path = row[5]
        us_path = row[6]

        # 3.根据类型拿到对应路径
        target_path = ""
        if audio_type == "uk":
            target_path = uk_path
        elif audio_type == "us":
            target_path = us_path
        else:
            return {"code":0, "msg":"音频类型只能为uk/us"}

        # 4.校验路径非空、文件真实存在
        if not target_path or len(str(target_path).strip()) == 0:
            return {"code": 0, "msg": "该单词无对应音频路径"}
        
        real_path = str(target_path).strip()
        if not os.path.isfile(real_path):
            return {"code": 0, "msg": f"音频文件不存在：{real_path}"}
        
        return {
            "code": 1,
            "file_path": real_path
        }
    

    @classmethod
    async def export_table_to_excel(cls, table_name: str):
        """
        导出指定单词表全量数据为Excel二进制流
        :param table_name: 目标单词表名称
        :return: 成功{code:1, excel_bytes:BytesIO, filename:str} 失败{code:0, msg:str}
        """
        # 1. 校验单词表是否真实存在
        check_table_sql = sql.check_name_vocabulary_exists_sql()
        table_check_rows = await db.execute_sql(check_table_sql, (table_name,))
        if table_check_rows[0][0] == 0:
            return {"code": 0, "msg": f"单词表【{table_name}】不存在"}

        # 2. 查询全表数据
        try:
            full_data_sql = sql.get_word_table_full_data_sql(table_name)
            rows = await db.execute_sql(full_data_sql)
        except Exception as e:
            return {"code": 0, "msg": f"查询单词表数据失败：{str(e)}"}

        # 3. 无数据兜底
        if not rows or len(rows) == 0:
            return {"code": 0, "msg": "该单词表无数据可导出"}

        # 固定单词表字段顺序（和建表结构完全匹配，彻底抛弃db.cursor读取）
        field_names = [
            "word",
            "translations",
            "exchanges",
            "examples",
            "phonetic",
            "uk_audio_path",
            "us_audio_path",
            "createTime"
        ]

        # 4. 创建Excel工作簿，写入数据
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{table_name}_单词表"

        # 写入表头（加粗样式）
        header_font = openpyxl.styles.Font(bold=True)
        for col_idx, field_name in enumerate(field_names, 1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = header_font
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 28

        # 写入数据行，空值转空字符串
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                val = str(cell_value) if cell_value is not None else ""
                ws.cell(row=row_idx, column=col_idx, value=val)

        # 5. 生成Excel二进制流
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return {
            "code": 1,
            "excel_bytes": excel_buffer,
            "filename": f"{table_name}_单词表全量数据.xlsx"
        }
    

    








