
import json
import openpyxl
from io import BytesIO

from database.sql import sql
from database.db import async_database as db



class paper:

    @classmethod
    async def add_paper_table(cls, table_name: str):
        #创建试卷表（含paper_table记录+具体试卷题目表）
        if not table_name or len(table_name.strip()) < 1:
            return {"code": 0, "msg": "试卷名称不能为空"}
        clean_name = table_name.strip()

        # 第一步：查重
        check_sql = sql.check_name_paper_exists_sql()
        rows = await db.execute_sql(check_sql, (clean_name,))
        count = rows[0][0]
        if count > 0:
            return {"code": 0, "msg": "该试卷名称已存在，请更换名称"}

        # 第二步：创建试卷表+插入paper_table记录
        sql_str = sql.create_paper_insert_paper_sql(clean_name)
        try:
            affect_rows = await db.execute_sql(sql_str, (clean_name, clean_name, clean_name))
        except Exception as e:
            print("数据库异常", repr(e))
            return {"code": 0, "msg": f"创建失败：{str(e)}"}

        if affect_rows >= 0:
            return {"code": 1, "msg": "试卷创建成功"}
        else:
            return {"code": 0, "msg": "创建执行无生效行数"}


    @classmethod
    async def get_all_paper_tables(cls):
        # 获取所有试卷列表
        sql_str = sql.get_all_paper_list_sql()
        rows = await db.execute_sql(sql_str)
        table_list = []
        for row in rows:
            table_list.append({
                "paperId": row[0],
                "paperName": row[1],
                "createTime": row[2]
            })
        return {"code": 1, "data": table_list}


    @classmethod
    async def search_paper_tables(cls, keyword: str):
        # 搜索试卷（内存过滤，避免SQL注入）
        all_tables_res = await cls.get_all_paper_tables()
        all_tables = all_tables_res.get("data", [])

        if not keyword or keyword.strip() == "":
            return all_tables_res

        filter_list = [
            table for table in all_tables
            if keyword.lower() in table["paperName"].lower()
        ]
        return {"code": 1, "data": filter_list}


    @classmethod
    async def delete_paper_table(cls, table_name: str):
        # 删除试卷（含paper_table记录+题目表）
        if not table_name or len(table_name.strip()) < 1:
            return {"code": 0, "msg": "试卷名称不能为空"}
        clean_name = table_name.strip()
        sql_str = sql.delete_paper_sql(clean_name)
        affect_rows = await db.execute_sql(sql_str, (clean_name,))

        if affect_rows >= 0:
            return {"code": 1, "msg": "试卷及内部题目数据已全部删除"}
        else:
            return {"code": 0, "msg": "删除试卷操作失败"}


    @classmethod
    async def import_questions_from_json(cls, table_name: str, json_file_bytes):
        # 从JSON导入题目到指定试卷
        BATCH_SIZE = 500

        # 1. 校验试卷是否存在
        check_table_sql = sql.check_name_paper_exists_sql()
        table_exist_rows = await db.execute_sql(check_table_sql, (table_name,))
        if table_exist_rows[0][0] == 0:
            return {"code": 0, "msg": f"试卷【{table_name}】不存在"}

        # 2. 解析JSON
        try:
            json_text = json_file_bytes.decode("utf-8")
            json_data = json.loads(json_text)
        except UnicodeDecodeError:
            return {"code": 0, "msg": "文件编码错误，请使用UTF-8格式JSON文件"}
        except json.JSONDecodeError as e:
            return {"code": 0, "msg": f"JSON格式解析失败：{str(e)}"}

        valid_question_list = []
        skip_count = 0

        # 3. 遍历清洗数据
        for key, question_item in json_data.items():
            if not isinstance(question_item, dict) or len(question_item) == 0:
                skip_count += 1
                continue

            # 基础字段
            question = question_item.get("question")
            answer = question_item.get("answer", "")
            objective = question_item.get("objective", False)
            # 新增分值字段，缺失给默认0.0
            score = float(question_item.get("score", 0.0))
            totalScore = float(question_item.get("totalScore", 0.0))

            # 必填校验
            if not question:
                skip_count += 1
                continue

            # 类型转换
            objective_bool = bool(objective) if isinstance(objective, (bool, int)) else False

            # 元组新增 score、totalScore 两个参数
            valid_question_list.append((
                question, answer, objective_bool, score, totalScore
            ))

        total_valid = len(valid_question_list)
        if total_valid == 0:
            return {"code": 0, "msg": f"无合法可导入题目，跳过数量：{skip_count}"}

        total_insert = 0
        # 4. 批量插入
        try:
            for start in range(0, total_valid, BATCH_SIZE):
                batch_items = valid_question_list[start:start + BATCH_SIZE]
                batch_len = len(batch_items)
                batch_sql = sql.batch_insert_question_sql(table_name, batch_len)

                flat_params = []
                for item in batch_items:
                    flat_params.extend(item)

                affect = await db.execute_sql(batch_sql, tuple(flat_params))
                total_insert += affect
        except Exception as e:
            print("批量插入题目异常", repr(e))
            return {
                "code": 0,
                "msg": f"数据库插入中断，已成功写入{total_insert}条，失败原因：{str(e)}",
                "count": total_insert
            }

        return {
            "code": 1,
            "msg": f"导入完成！成功{total_insert}条，跳过非法数据{skip_count}",
            "count": total_insert
        }


    @classmethod
    async def get_paper_all_questions(cls, table_name: str):
        # 获取指定试卷所有题目
        # 校验试卷存在
        check_sql = sql.check_name_paper_exists_sql()
        row = await db.execute_sql(check_sql, (table_name,))
        if row[0][0] == 0:
            return {"code": 0, "msg": f"试卷{table_name}不存在"}

        get_questions_sql = sql.get_paper_all_questions_sql(table_name)
        rows = await db.execute_sql(get_questions_sql)
        question_list = []
        for r in rows:
            question_list.append({
                "questionId": r[0],
                "question": r[1],
                "answer": r[2],
                "objective": bool(r[3]) if r[3] is not None else False
            })
        return {
            "code": 1,
            "msg": "查询成功",
            "data": question_list
        }


    @classmethod
    async def get_single_question_info(cls, table_name: str, question_id: int):
        # 获取单道题目详情
        # 校验试卷存在
        check_sql = sql.check_name_paper_exists_sql()
        row = await db.execute_sql(check_sql, (table_name,))
        if row[0][0] == 0:
            return {"code": 0, "msg": f"试卷{table_name}不存在"}

        get_question_sql = sql.get_single_question_detail_sql(table_name)
        rows = await db.execute_sql(get_question_sql, (question_id,))
        if not rows:
            return {"code": 0, "msg": "该题目不存在"}
        r = rows[0]
        data = {
            "questionId": r[0],
            "question": r[1],
            "answer": r[2],
            "objective": bool(r[3]) if r[3] is not None else False
        }
        return {
            "code": 1,
            "data": data
        }


    @classmethod
    async def delete_single_question(cls, table_name: str, question_id: int):
        # 删除单道题目
        # 校验试卷存在
        check_sql = sql.check_name_paper_exists_sql()
        row = await db.execute_sql(check_sql, (table_name,))
        if row[0][0] == 0:
            return {"code": 0, "msg": f"试卷{table_name}不存在"}

        # 执行删除
        delete_sql = sql.delete_single_question_sql(table_name)
        affect_rows = await db.execute_sql(delete_sql, (question_id,))
        if affect_rows > 0:
            return {"code": 1, "msg": "题目删除成功"}
        else:
            return {"code": 0, "msg": "题目删除失败（题目ID不存在或数据库异常）"}


    @classmethod
    async def update_question(cls, table_name: str, question_id: int, question: str, answer: str, objective: bool):
        # 更新题目内容
        # 校验试卷存在
        check_sql = sql.check_name_paper_exists_sql()
        row = await db.execute_sql(check_sql, (table_name,))
        if row[0][0] == 0:
            return {"code": 0, "msg": f"试卷{table_name}不存在"}

        # 必填校验
        if not question:
            return {"code": 0, "msg": "题目内容不能为空"}

        # 执行更新
        update_sql = sql.update_question_sql(table_name)
        affect_rows = await db.execute_sql(update_sql, (question, answer, objective, question_id))
        if affect_rows > 0:
            return {"code": 1, "msg": "题目更新成功"}
        else:
            return {"code": 0, "msg": "题目更新失败（题目ID不存在或数据无变更）"}


    @classmethod
    async def export_paper_to_excel(cls, table_name: str):
        # 导出试卷为Excel二进制流
        # 1. 校验试卷存在
        check_table_sql = sql.check_name_paper_exists_sql()
        table_check_rows = await db.execute_sql(check_table_sql, (table_name,))
        if table_check_rows[0][0] == 0:
            return {"code": 0, "msg": f"试卷【{table_name}】不存在"}

        # 2. 查询全表题目
        try:
            full_data_sql = sql.get_paper_all_questions_sql(table_name)
            rows = await db.execute_sql(full_data_sql)
        except Exception as e:
            return {"code": 0, "msg": f"查询试卷数据失败：{str(e)}"}

        # 3. 无数据兜底
        if not rows or len(rows) == 0:
            return {"code": 0, "msg": "该试卷无题目可导出"}

        # 固定字段顺序
        field_names = ["questionId", "question", "answer", "objective"]

        # 4. 创建Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{table_name}_试卷"

        # 写入表头（加粗）
        header_font = openpyxl.styles.Font(bold=True)
        for col_idx, field_name in enumerate(field_names, 1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = header_font
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 35

        # 写入数据行
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                # 布尔值转中文
                if col_idx == 4:
                    val = "客观题" if cell_value else "主观题"
                else:
                    val = str(cell_value) if cell_value is not None else ""
                ws.cell(row=row_idx, column=col_idx, value=val)

        # 5. 生成二进制流
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return {
            "code": 1,
            "excel_bytes": excel_buffer,
            "filename": f"{table_name}_试卷全量题目.xlsx"
        }





